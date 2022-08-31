import time
# import openpyxl module
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By


browser = webdriver.Chrome(executable_path="./chromedriver")
#browser.get('http://10.115.57.132:80/maPage.html')
browser.get('http://localhost:8080/maPage.html')

testNom = 'toto'

def testMajuscule(valeur):
    elem = browser.find_element(By.NAME, 'nom')  # Find the search box
    elem.clear()
    elem.send_keys( valeur )

    bouton = browser.find_element(By.NAME, 'bouton')
    bouton.click()

    div = browser.find_element(By.NAME, 'resultat')
    #result = div.getText()
    result = div. get_attribute( 'innerHTML' )


    if result == valeur.upper():
        print( 'OK' )
    else: 
        print( 'KO' )

def testCalcule(ligne):

    # Python program to read an excel file
    
    # Give the location of the file
    path = "C:\\xampp\\htdocs\\casdetest.xlsx"
    
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    
    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.
    
    # Note: The first row or
    # column integer is 1, not 0.
    
    # Cell object is created by using
    # sheet object's cell() method.

    cell_obj = sheet_obj.cell(row = ligne+1, column = 1)

    elem = browser.find_element(By.NAME, 'saisie')  # Find the search box
    elem.clear()
    elem.send_keys( cell_obj.value )

    bouton = browser.find_element(By.NAME, 'go')
    bouton.click()

    div = browser.find_element(By.NAME, 'res')
    #result = div.getText()
    result = div. get_attribute( 'innerHTML' )
    print(result)
        
    # Print value of cell object
    # using the value attribute
    print(cell_obj.value)
    cell_obj = sheet_obj.cell(row = ligne+1, column = 2)
    print(cell_obj.value)

    


    cell_res = sheet_obj.cell(row = ligne+1, column = 3)
    if int(result) == cell_obj.value:
        print( 'OK' )
        cell_res.value = "OK"
    else: 
        print( 'KO' )
        cell_res.value = "KO"

    wb_obj.save("C:\\xampp\\htdocs\\casdetest.xlsx")

testMajuscule('toto')
testMajuscule('tata')
testMajuscule('1234')
testMajuscule('t#@')

for i in range(3):
    testCalcule(i)


time.sleep(1)
browser.quit()